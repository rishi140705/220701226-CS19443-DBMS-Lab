import customtkinter
from tkinter import *
from customtkinter import CTkImage
from tkinter import messagebox
from PIL import Image
import openpyxl as xl
from datetime import date
import mysql.connector

app = customtkinter.CTk()
app.geometry("800x450")
app.config(bg="#000000")
app.title("Supermarket")

img1 = CTkImage(Image.open(r'chocolate.png'), size=(100, 100))
img2 = CTkImage(Image.open(r'coke.png'), size=(100, 100))
img3 = CTkImage(Image.open(r'milk.png'), size=(100, 100))
img4 = CTkImage(Image.open(r'oill.png'), size=(100, 100))

font1 = ('arial', 15, 'bold')

Variable1 = IntVar()
Variable2 = IntVar()
Variable3 = IntVar()
Variable4 = IntVar()
Variable5 = 0
Variable6 = ''

price_list = [20, 10, 20, 30]

def get_db_connection():
    return mysql.connector.connect(
        host="localhost", 
        user="root",
        password="Changeme@123",
        database="supermarket"
    )

def cancel():
    Variable1.set(0)
    Variable2.set(0)
    Variable3.set(0)
    Variable4.set(0)

def pay():
    global Variable5, Variable6
    if name_entry.get() == '':
        messagebox.showerror(title="Error", message="Please enter your name")
    else:
        Variable5 = Variable1.get() * price_list[0] + Variable2.get() * price_list[1] + Variable3.get() * price_list[2] + Variable4.get() * price_list[3]

        total_label = customtkinter.CTkLabel(app, text=Variable5, font=font1, text_color="#000000", fg_color="#FFFFFF", width=300)
        total_label.place(x=628, y=150)

        Variable6 = date.today()

        date_label = customtkinter.CTkLabel(app, text=Variable6, font=font1, text_color="#000000", fg_color="#FFFFFF", width=300)
        date_label.place(x=628, y=250)

def confirm():
    global Variable5, Variable6
    file = xl.load_workbook('customers.xlsx')
    sheet = file["Sheet1"]
    sheet.cell(column=1, row=sheet.max_row + 1, value=name_entry.get())
    sheet.cell(column=2, row=sheet.max_row, value=Variable5)
    sheet.cell(column=3, row=sheet.max_row, value=Variable6)
    file.save('customers.xlsx')

    
    db = get_db_connection()
    cursor = db.cursor() 
    sql = "INSERT INTO customers (name, total_price, bill_date) VALUES (%s, %s, %s)"
    values = (name_entry.get(), Variable5, Variable6)
    cursor.execute(sql, values)
    db.commit()
    cursor.close()
    db.close()

    messagebox.showinfo(title="Success", message="Data has been saved to Excel and MySQL database")

def clear_data():
    db = get_db_connection()
    cursor = db.cursor()
    cursor.execute("TRUNCATE TABLE customers")
    db.commit()
    cursor.close()
    db.close()
    
    file = xl.load_workbook('customers.xlsx')
    sheet = file["Sheet1"]
    sheet.delete_rows(2, sheet.max_row)
    file.save('customers.xlsx')

    messagebox.showinfo(title="Success", message="All data has been cleared from MySQL database and Excel file")

button1 = customtkinter.CTkButton(app, text="chocolate", font=font1, fg_color="#0e1d54", hover_color="#0e1d54", text_color="#FFFFFF", border_width=1, width=150, height=200, image=img1, compound=TOP)
button1.place(x=30, y=20)

button2 = customtkinter.CTkButton(app, text="coke", font=font1, fg_color="#0e1d54", hover_color="#0e1d54", text_color="#FFFFFF", border_width=1, width=150, height=200, image=img2, compound=TOP)
button2.place(x=300, y=20)

button3 = customtkinter.CTkButton(app, text="milk", font=font1, fg_color="#0e1d54", hover_color="#0e1d54", text_color="#FFFFFF", border_width=1, width=150, height=200, image=img3, compound=TOP)
button3.place(x=30, y=270)

button4 = customtkinter.CTkButton(app, text="oil", font=font1, fg_color="#0e1d54", hover_color="#0e1d54", text_color="#FFFFFF", border_width=1, width=150, height=200, image=img4, compound=TOP)
button4.place(x=300, y=270)

sp1 = Spinbox(app, from_=0, to=10, font=font1, width=8, background="#5c99ad", justify=CENTER, textvariable=Variable1)
sp1.place(x=75, y=280)

sp2 = Spinbox(app, from_=0, to=10, font=font1, width=8, background="#5c99ad", justify=CENTER, textvariable=Variable2)
sp2.place(x=415, y=280)

sp3 = Spinbox(app, from_=0, to=10, font=font1, width=8, background="#5c99ad", justify=CENTER, textvariable=Variable3)
sp3.place(x=75, y=593)

sp4 = Spinbox(app, from_=0, to=10, font=font1, width=8, background="#5c99ad", justify=CENTER, textvariable=Variable4)
sp4.place(x=415, y=593)

name_label = customtkinter.CTkLabel(app, text="Customer Name:", font=font1, text_color="#FFFFFF")
name_label.place(x=500, y=50)

t_Price = customtkinter.CTkLabel(app, text="Total Price:", font=font1, text_color="#FFFFFF")
t_Price.place(x=537.5, y=150)

t_date = customtkinter.CTkLabel(app, text="Bill Date:", font=font1, text_color="#FFFFFF")
t_date.place(x=557, y=250)

name_entry = customtkinter.CTkEntry(app, font=font1, text_color="#000000", fg_color="#FFFFFF", border_color="#000000", width=300)
name_entry.place(x=628, y=50)

pay_button = customtkinter.CTkButton(app, command=pay, text="Pay Bill", fg_color="#48158a", hover_color="#48158a", text_color="#FFFFFF", font=font1)
pay_button.place(x=650, y=300)

cancel_button = customtkinter.CTkButton(app, command=cancel, text="Cancel Bill", fg_color="#cf840c", hover_color="#cf840c", text_color="#FFFFFF", font=font1)
cancel_button.place(x=800, y=300)

confirm_button = customtkinter.CTkButton(app, command=confirm, text="Confirm Payment", fg_color="#11b81e", hover_color="#11b81e", text_color="#FFFFFF", font=font1)
confirm_button.place(x=725, y=350)

clear_button = customtkinter.CTkButton(app, command=clear_data, text="Clear Data", fg_color="#ff0000", hover_color="#ff0000", text_color="#FFFFFF", font=font1)
clear_button.place(x=725, y=400)

app.mainloop()